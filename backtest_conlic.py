# -*- coding: utf-8 -*-
"""
backtest_conlic.py — A PROVA pra diretoria (decisao do conselho 2026-07-04).

Compara o boletim do ConLicitacao (o .xlsx que o Igor recebe) com o que o NOSSO
radar coletou no mesmo periodo. Responde a pergunta que a diretoria faz:
"quantas licitacoes de concreto/brita em Pregao Eletronico o radar gratis pegou,
das que o ConLicitacao (pago) listou?"

NAO cancela nada, NAO grava nada — so imprime o placar + a lista dos que faltaram
com o motivo provavel (fora do raio / nao coletado). Rode em paralelo por semanas.

Uso:
  python backtest_conlic.py "caminho/do/boletim.xlsx"
  (ou defina CONLIC_XLSX=... no ambiente)

Precisa de openpyxl (ja e dependencia do projeto via pandas). O radar coletado e
lido de docs/dados.json (historico acumulado do site).
"""
import sys, os, re, json, unicodedata
try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl: pip install openpyxl"); sys.exit(1)


def _n(s):
    return re.sub(r'\s+', ' ', unicodedata.normalize('NFKD', str(s or ''))
                  .encode('ascii', 'ignore').decode().lower()).strip()

# ---- relevancia concreto/brita (mesma logica do radar.py: exclui asfalto/pre-moldado) ----
KW3 = ("concreto usinado", "concreto pre-misturado", "concreto pre misturado", "concreto dosado",
       "central dosadora", "fornecimento de concreto", "concreto bombeado", "concreto bombeavel",
       "concreto fck", "concreto estrutural", "concreto convencional",
       "brita", "britas", "brita graduada", "bgs", "pedra britada", "pedras britadas", "pedrisco",
       "po de pedra", "bica corrida", "rachao", "cascalho", "agregado graudo", "pedra de mao", "seixo")
KW2_CONC = ("concreto", "concretagem", "concreto armado")
KW2_BRITA = ("agregado", "agregados")
CONTEXTO = ("fck", "mpa", "m3", "metro cubico", "usina", "usinado", "central", "dosado", "bombeado", "betoneira")
EXCL = ("tubo de concreto", "manilha", "aduela", "poste de concreto", "bloco de concreto", "bloco estrutural",
        "bloquete", "artefato de concreto", "pre-moldado", "pre moldado", "premoldado", "pre-fabricado",
        "piso intertravado", "paver", "lajota", "meio-fio", "meio fio", "sarjeta", "cimento", "argamassa",
        "pavimentacao asfaltica", "asfalto", "cbuq", "concreto asfaltico", "concreto betuminoso")
HARD_EXCL = ("asfalt", "cbuq", "betumin")

def score(texto):
    t = _n(texto)
    if any(e in t for e in HARD_EXCL): return 0
    if any(k in t for k in KW3): return 3
    if any(e in t for e in EXCL): return 0
    if any(k in t for k in KW2_BRITA): return 2
    if any(k in t for k in KW2_CONC) and any(c in t for c in CONTEXTO): return 2
    return 0

def rel(t):
    return score(t) >= 2

def is_pe(edital, status):
    """PE pelo prefixo do numero (ex 'PE/0006/2025') ou pela ausencia de marca de outra modalidade."""
    e = _n(edital)
    if e.startswith('pe') or 'pregao eletronico' in _n(status) or 'eletronico' in _n(status):
        return True
    # se marca explicitamente outra modalidade, fora
    if any(m in e for m in ('cr/', 'dl/', 'pr/', 'sm/', 'concorrencia', 'dispensa', 'presencial')):
        return False
    return True   # default: boletim ConLic ja costuma vir de PE; tratamos como PE salvo marca contraria

_COLMAP = {
    'objeto': ('objeto', 'descricao', 'objeto da licitacao'),
    'edital': ('edital', 'numero', 'numero do edital', 'n edital'),
    'orgao': ('orgao', 'orgao/entidade', 'entidade', 'comprador'),
    'cidade': ('cidade', 'municipio', 'cidade/uf', 'municipio/uf'),
    'uf': ('estado', 'uf', 'sigla uf'),
    'valor': ('valor estimado', 'valor', 'valor estimado (r$)'),
    'nc': ('numero conlicitacao', 'nc', 'controle'),
    'status': ('situacao', 'status'),
}

def _pega(row_norm, canon):
    for v in _COLMAP.get(canon, ()):
        if _n(v) in row_norm and row_norm[_n(v)]:
            return str(row_norm[_n(v)]).strip()
    return ''

def ler_conlic(caminho):
    wb = load_workbook(caminho, read_only=True, data_only=True)
    linhas = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        hdr = [_n(h) for h in rows[0]]
        for r in rows[1:]:
            if not any(r):
                continue
            rn = {hdr[i]: r[i] for i in range(min(len(hdr), len(r))) if hdr[i]}
            linhas.append({
                'objeto': _pega(rn, 'objeto'), 'edital': _pega(rn, 'edital'),
                'orgao': _pega(rn, 'orgao'), 'cidade': _pega(rn, 'cidade'),
                'uf': _pega(rn, 'uf'), 'valor': _pega(rn, 'valor'),
                'nc': _pega(rn, 'nc'), 'status': _pega(rn, 'status'), 'aba': ws.title,
            })
    return linhas

def _muni(s):
    # 'Muriae/MG' ou 'Muriae - MG' -> 'muriae'
    s = _n(s)
    s = re.sub(r'\s*[/-]\s*[a-z]{2}\s*$', '', s)
    return s.strip()

def _numcore(s):
    return re.sub(r'[^0-9]', '', _n(s))   # so os digitos do numero do edital

def main():
    caminho = (sys.argv[1] if len(sys.argv) > 1 else os.environ.get('CONLIC_XLSX', '')).strip('"')
    if not caminho or not os.path.exists(caminho):
        print("Uso: python backtest_conlic.py \"caminho/boletim.xlsx\"  (ou CONLIC_XLSX=...)")
        print("Arquivo nao encontrado:", repr(caminho)); return

    conlic = ler_conlic(caminho)
    print("Linhas no boletim ConLic:", len(conlic))
    alvo = [c for c in conlic if rel(c['objeto']) and is_pe(c['edital'], c['status'])]
    print("-> destas, PE de concreto/brita (nosso escopo):", len(alvo))
    if not alvo:
        print("Nenhuma linha PE concreto/brita no boletim. Confira as colunas do xlsx.")
        # ajuda a diagnosticar colunas
        if conlic:
            print("Colunas lidas da 1a linha:", {k: (v[:30] if isinstance(v, str) else v) for k, v in conlic[0].items()})
        return

    # radar coletado (historico acumulado do site)
    dj = os.path.join('docs', 'dados.json')
    dados = json.load(open(dj, encoding='utf-8')) if os.path.exists(dj) else []
    idx_num = {}
    idx_mun = {}
    for r in dados:
        mun = _muni(r.get('MUNICIPIO', ''))
        nc = _numcore(r.get('NUMERO', ''))
        if mun and nc:
            idx_num[(mun, nc)] = r
        idx_mun.setdefault(mun, []).append(r)

    def achou(c):
        mun = _muni(c['cidade']); nc = _numcore(c['edital'])
        if mun and nc and (mun, nc) in idx_num:
            return True
        # fallback: mesmo municipio + sobreposicao forte de objeto (concreto/brita ja garantido)
        for r in idx_mun.get(mun, []):
            if rel(r.get('OBJETO', '')):
                return True
        return False

    pego = [c for c in alvo if achou(c)]
    faltou = [c for c in alvo if not achou(c)]
    n = len(alvo)
    print("\n================= PLACAR (PILOTO) =================")
    print("Radar pegou: %d de %d editais PE concreto/brita" % (len(pego), n))
    print("Faltaram:    %d (conferir se sao FORA DO RAIO = ok, ou perda real)" % len(faltou))
    print("AMOSTRA PEQUENA — nao afirmar '100%'. Diga: '%d/%d no piloto, seguimos medindo'." % (len(pego), n))
    print("==================================================")
    if faltou:
        print("\n--- FALTARAM (conferir: fora do raio? fonte? atraso?) ---")
        for c in faltou[:40]:
            print("  [%s] %-22s | %s | %s" % (c['uf'], (c['cidade'] or '?')[:22],
                                              (c['edital'] or '?')[:14], (c['objeto'] or '')[:60]))
    print("\nObs: 'faltou' pode ser edital FORA DO RAIO (concreto>70km/brita>500km) — que o radar")
    print("descarta de proposito. Cruze a cidade com as filiais antes de considerar perda real.")

if __name__ == '__main__':
    main()
